# ------Import all required libraries------ #
import banner1
import colors
import random
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import os
# import openpyxl
import time
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side
import datetime
from pathlib import Path

if os.name == "nt":
    os.system("mode 600")
else:
    os.system("wmctrl -r :ACTIVE: -b add,fullscreen")

date = f"{datetime.datetime.now():%Y-%m-%d}"
if os.name == "nt":
    os.system("cls")
else:
    os.system("clear")

foo = [banner1.banner1]
print(colors.CWHITE + random.choice(foo))
time.sleep(1)
print("""
Welcome to GPA calculator for UAF purely build using PYTHON.
""")
print("""
1.Enter your full AG number.
2.Enter your current semester.
3.Enter number of courses in each semester.
""")
time.sleep(3)
move_on = input(colors.CWHITE + "Press any key to continue")
if os.name == "nt":
    os.system("cls")
else:
    os.system("clear")
# ------Get Ag No. and start web driver ------ #


USERNAME = input("Enter your AG number (e.g. 2019-ag-6737): ")
last_semester = int(input("Your last semester was(1-8) ? "))
if last_semester != 0:
    for i in range(1, last_semester+1):
        i = str(i)
        globals()['number_of_subjects_%s' % i] = int(input("Number of subjects/courses in semester " + i + " ? "))
elif last_semester == 0:
    print(colors.CRED + """           
                
        Masti Kr rea wan
        Tu bari haram mout marna
    """)
    print("")
    exit()
path = Path.cwd()
print("Connecting to LMS......")
options = Options()
# -- options.binary_location = "C:\\Program Files\\Google\\Chrome Dev\\Application\\chrome.exe"  -- #
# -- //you can provide manual path to chrome app using above statement -- #
options.add_argument("headless")

options.add_argument('--log-level=3')
if os.name == "nt":
    driver = webdriver.Chrome(options=options, executable_path=r'{}/chromedriver.exe'.format(path))
else:
    driver = webdriver.Chrome(options=options, executable_path=r'{}/chromedriver'.format(path))
driver.get("http://lms.uaf.edu.pk/login/index.php")

user = driver.find_element_by_id('REG')
user.send_keys(USERNAME)

print("Entering Values")

button = driver.find_element_by_xpath("//input[@value='Result']")
button.click()

print("Extracting Result")
student_ag = driver.find_element_by_xpath('/html/body/table[1]/tbody/tr[1]/td[2]').text
student_name = driver.find_element_by_xpath('/html/body/table[1]/tbody/tr[2]/td[2]').text
print("Student Name: {}".format(student_name))
print("Student AG: {}".format(student_ag))
row_count = len(driver.find_elements_by_xpath('/html/body/table[2]/tbody/tr'))
first_col_count = len(driver.find_elements_by_xpath('/html/body/table[2]/tbody/tr[2]/td'))
# second_col_count = len(driver.find_elements_by_xpath('/html/body/table[2]/tbody/tr[2]/td'))

# print("Number of columns : {}".format(second_col_count))
first_part = '/html/body/table[2]/tbody/tr['
second_part = ']/td['
third_part = ']'
print("Generating SpreadSheet")


# ------Generate spreadsheet------ #

total_num = []
for n in range(2, row_count + 1):
    for m in range(1, first_col_count + 1):
        final_path = first_part + str(n) + second_part + str(m) + third_part
        table_data = driver.find_element_by_xpath(final_path).text
        fname = USERNAME + ".xlsx"
        if os.path.exists(fname):
            wb = load_workbook(fname)
            ws = wb['Sheet']
        else:
            wb = Workbook()
            ws = wb.active
            # ------Style Spreadsheet------ #
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        no_border = Border(left=Side(style='none'),
                           right=Side(style='none'),
                           top=Side(style='none'),
                           bottom=Side(style='none'))

        ws.cell(row=n + 3, column=m).value = table_data
        ws.cell(row=n + 3, column=m).border = thin_border
        ws.cell(row=n + 3, column=1).border = no_border

        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 24
        ws.column_dimensions['D'].width = 16
        ws.column_dimensions['E'].width = 45
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['H'].width = 14
        ws.column_dimensions['J'].width = 9.5
        ws['F2'] = "Printed On :"
        ws['F2'].font = Font(bold=True, size=10, name="Google Sans")
        ws['H2'] = date
        ws['H2'].font = Font(bold=True, size=10, name="Google Sans")
        ws['B2'] = "Registration # :"
        ws['B2'].font = Font(bold=True, size=16, name="Google Sans")
        ws.cell(row=2, column=2).border = thin_border
        ws['C2'] = student_ag
        ws['C2'].font = Font(bold=True, size=16, name="Google Sans")
        ws.cell(row=2, column=3).border = thin_border
        ws['B3'] = "Student Name: "
        ws['B3'].font = Font(bold=True, size=16, name="Google Sans")
        ws.cell(row=3, column=2).border = thin_border
        ws.cell(row=3, column=2).alignment = Alignment(vertical='center')
        ws['C3'] = student_name
        ws['C3'].font = Font(bold=True, size=16, name="Google Sans")
        ws['C3'].alignment = Alignment(wrap_text=True)
        ws.cell(row=3, column=3).border = thin_border
        ws['A4'] = 'Sr.'
        ws['A4'].font = Font(bold=True, name="Google Sans")
        ws['B4'] = 'Semester'
        ws['B4'].font = Font(bold=True, name="Google Sans")
        ws.cell(row=4, column=2).border = thin_border
        ws['C4'] = 'Teacher Name'
        ws['C4'].font = Font(bold=True, name="Google Sans")
        ws.cell(row=4, column=3).border = thin_border
        ws['D4'] = 'Course Code'
        ws['D4'].font = Font(bold=True, name="Google Sans")
        ws.cell(row=4, column=4).border = thin_border
        ws['E4'] = 'Course Title'
        ws['E4'].font = Font(bold=True, name="Google Sans")
        ws.cell(row=4, column=5).border = thin_border
        ws['F4'] = 'Credit Hours'
        ws['F4'].font = Font(bold=True, name="Google Sans")
        ws.cell(row=4, column=6).border = thin_border
        ws['G4'] = 'Mid'
        ws['G4'].font = Font(bold=True, name="Google Sans")
        ws.cell(row=4, column=7).border = thin_border
        ws['H4'] = 'Assignment'
        ws['H4'].font = Font(bold=True, name="Google Sans")
        ws.cell(row=4, column=8).border = thin_border
        ws['I4'] = 'Final'
        ws['I4'].font = Font(bold=True, name="Google Sans")
        ws.cell(row=4, column=9).border = thin_border
        ws['J4'] = 'Practical'
        ws['J4'].font = Font(bold=True, name="Google Sans")
        ws.cell(row=4, column=10).border = thin_border
        ws['K4'] = 'Total'
        ws['K4'].font = Font(bold=True, name="Google Sans")
        ws.cell(row=4, column=11).border = thin_border
        ws['L4'] = 'Grade'
        ws['L4'].font = Font(bold=True, name="Google Sans")
        ws.cell(row=4, column=12).border = thin_border
        wb.save(fname)
