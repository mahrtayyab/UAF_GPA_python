import time
import LMS
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment
import subprocess
import os

from pathlib import Path

# -- Quality Points --#

qp21 = 1
qp22 = 1.5
qp23 = 2
qp24 = 2.33
qp25 = 2.67
qp26 = 3
qp27 = 3.33
qp28 = 3.67
qp29 = 4
qp20 = 0

qp61 = 3
qp62 = 3.5
qp63 = 4
qp64 = 4.5
qp65 = 5
qp66 = 5.5
qp67 = 6
qp68 = 6.33
qp69 = 6.67
qp70 = 7
qp71 = 7.33
qp72 = 7.67
qp73 = 8
qp74 = 8.33
qp75 = 8.67
qp76 = 9
qp77 = 9.33
qp78 = 9.67
qp79 = 10
qp80 = 10.33
qp81 = 10.67
qp82 = 11
qp83 = 11.33
qp84 = 11.67
qp85 = 12

qp41 = 2
qp42 = 2.5
qp43 = 3
qp44 = 3.5
qp45 = 4
qp46 = 4.33
qp47 = 4.67
qp48 = 5
qp49 = 5.33
qp50 = 5.67
qp51 = 6
qp52 = 6.33
qp53 = 6.67
qp54 = 7
qp55 = 7.33
qp56 = 7.67
qp57 = 8

qp86 = 4
qp87 = 4.5
qp88 = 5
qp89 = 5.5
qp90 = 6
qp91 = 6.5
qp92 = 7
qp93 = 7.5
qp94 = 8
qp95 = 8.33
qp96 = 8.67
qp97 = 9
qp98 = 9.33
qp99 = 9.67
qp100 = 10
qp101 = 10.33
qp102 = 10.67
qp103 = 11
qp104 = 11.33
qp105 = 11.67
qp106 = 12
qp107 = 12.33
qp108 = 12.67
qp109 = 13
qp110 = 13.33
qp111 = 13.67
qp112 = 14
qp113 = 14.33
qp114 = 14.67
qp115 = 15
qp116 = 15.33
qp117 = 15.67
qp118 = 16

qp120 = 5
qp121 = 5.5
qp122 = 6
qp123 = 6.5
qp124 = 7
qp125 = 7.5
qp126 = 8
qp127 = 8.5
qp128 = 9
qp129 = 9.5
qp130 = 10
qp131 = 10.33
qp132 = 10.67
qp133 = 11
qp134 = 11.33
qp135 = 11.67
qp136 = 12
qp137 = 12.33
qp138 = 12.67
qp139 = 13
qp140 = 13.33
qp141 = 13.67
qp142 = 14
qp143 = 14.33
qp144 = 14.67
qp145 = 15
qp146 = 15.33
qp147 = 15.67
qp148 = 16
qp149 = 16.33
qp150 = 16.67
qp151 = 17
qp152 = 17.33
qp153 = 17.67
qp154 = 18
qp155 = 18.33
qp156 = 18.67
qp157 = 19
qp158 = 19.33
qp159 = 19.67
qp160 = 20

# -- Quality Points -- #

wb = load_workbook(LMS.fname)
ws = wb['Sheet']
cell_count = ws.max_row
# real_cell_count = cell_count+1
last_row = cell_count + 1


def doit():
    if a == 24 and h == 60:
        a1 = qp61
    elif a == 25 and h == 60:
        a1 = qp62
    elif a == 26 and h == 60:
        a1 = qp63
    elif a == 27 and h == 60:
        a1 = qp64
    elif a == 28 and h == 60:
        a1 = qp65
    elif a == 29 and h == 60:
        a1 = qp66
    elif a == 30 and h == 60:
        a1 = qp67
    elif a == 31 and h == 60:
        a1 = qp68
    elif a == 32 and h == 60:
        a1 = qp69
    elif a == 33 and h == 60:
        a1 = qp70
    elif a == 34 and h == 60:
        a1 = qp71
    elif a == 35 and h == 60:
        a1 = qp72
    elif a == 36 and h == 60:
        a1 = qp73
    elif a == 37 and h == 60:
        a1 = qp74
    elif a == 38 and h == 60:
        a1 = qp75
    elif (a == 39) and (h == 60):
        a1 = qp76
    elif (a == 40) and (h == 60):
        a1 = qp77
    elif (a == 41) and (h == 60):
        a1 = qp78
    elif (a == 42) and (h == 60):
        a1 = qp79
    elif (a == 43) and (h == 60):
        a1 = qp80
    elif (a == 44) and (h == 60):
        a1 = qp81
    elif (a == 45) and (h == 60):
        a1 = qp82
    elif (a == 46) and (h == 60):
        a1 = qp83
    elif (a == 47) and (h == 60):
        a1 = qp84
    elif (a >= 48) and (a <= 60) and (h == 60):
        a1 = qp85
    elif (a == 8) and (h == 20):
        a1 = qp21
    elif (a == 9) and (h == 20):
        a1 = qp22
    elif (a == 10) and (h == 20):
        a1 = qp23
    elif (a == 11) and (h == 20):
        a1 = qp24
    elif (a == 12) and (h == 20):
        a1 = qp25
    elif (a == 13) and (h == 20):
        a1 = qp26
    elif (a == 14) and (h == 20):
        a1 = qp27
    elif (a == 15) and (h == 20):
        a1 = qp28
    elif (a >= 16) and (a <= 20) and (h == 20):
        a1 = qp29
    elif (a == 16) and (h == 40):
        a1 = qp41
    elif (a == 17) and (h == 40):
        a1 = qp42
    elif (a == 18) and (h == 40):
        a1 = qp43
    elif (a == 19) and (h == 40):
        a1 = qp44
    elif (a == 20) and (h == 40):
        a1 = qp45
    elif (a == 21) and (h == 40):
        a1 = qp46
    elif (a == 22) and (h == 40):
        a1 = qp47
    elif (a == 23) and (h == 40):
        a1 = qp48
    elif (a == 24) and (h == 40):
        a1 = qp49
    elif (a == 25) and (h == 40):
        a1 = qp50
    elif (a == 26) and (h == 40):
        a1 = qp51
    elif (a == 27) and (h == 40):
        a1 = qp52
    elif (a == 28) and (h == 40):
        a1 = qp53
    elif (a == 29) and (h == 40):
        a1 = qp54
    elif (a == 30) and (h == 40):
        a1 = qp55
    elif (a == 31) and (h == 40):
        a1 = qp56
    elif (a >= 32) and (a <= 40) and (h == 40):
        a1 = qp57
    elif (a == 32) and (h == 80):
        a1 = qp86
    elif (a == 33) and (h == 80):
        a1 = qp87
    elif (a == 34) and (h == 80):
        a1 = qp88
    elif (a == 35) and (h == 80):
        a1 = qp89
    elif (a == 36) and (h == 80):
        a1 = qp90
    elif (a == 37) and (h == 80):
        a1 = qp91
    elif (a == 38) and (h == 80):
        a1 = qp92
    elif (a == 39) and (h == 80):
        a1 = qp93
    elif (a == 40) and (h == 80):
        a1 = qp94
    elif (a == 41) and (h == 80):
        a1 = qp95
    elif (a == 42) and (h == 80):
        a1 = qp96
    elif (a == 43) and (h == 80):
        a1 = qp97
    elif (a == 44) and (h == 80):
        a1 = qp98
    elif (a == 45) and (h == 80):
        a1 = qp99
    elif (a == 46) and (h == 80):
        a1 = qp100
    elif (a == 47) and (h == 80):
        a1 = qp101
    elif (a == 48) and (h == 80):
        a1 = qp102
    elif (a == 49) and (h == 80):
        a1 = qp103
    elif (a == 50) and (h == 80):
        a1 = qp104
    elif (a == 51) and (h == 80):
        a1 = qp105
    elif (a == 52) and (h == 80):
        a1 = qp106
    elif (a == 53) and (h == 80):
        a1 = qp107
    elif (a == 54) and (h == 80):
        a1 = qp108
    elif (a == 55) and (h == 80):
        a1 = qp109
    elif (a == 56) and (h == 80):
        a1 = qp110
    elif (a == 57) and (h == 80):
        a1 = qp111
    elif (a == 58) and (h == 80):
        a1 = qp112
    elif (a == 59) and (h == 80):
        a1 = qp113
    elif (a == 60) and (h == 80):
        a1 = qp114
    elif (a == 61) and (h == 80):
        a1 = qp115
    elif (a == 62) and (h == 80):
        a1 = qp116
    elif (a == 63) and (h == 80):
        a1 = qp117
    elif (a >= 64) and (a <= 80) and (h == 80):
        a1 = qp118
    elif (a == 40) and (h == 100):
        a1 = qp120
    elif (a == 41) and (h == 100):
        a1 = qp121
    elif (a == 42) and (h == 100):
        a1 = qp122
    elif (a == 43) and (h == 100):
        a1 = qp123
    elif (a == 44) and (h == 100):
        a1 = qp124
    elif (a == 45) and (h == 100):
        a1 = qp125
    elif (a == 46) and (h == 100):
        a1 = qp126
    elif (a == 47) and (h == 100):
        a1 = qp127
    elif (a == 48) and (h == 100):
        a1 = qp128
    elif (a == 49) and (h == 100):
        a1 = qp129
    elif (a == 50) and (h == 100):
        a1 = qp130
    elif (a == 51) and (h == 100):
        a1 = qp131
    elif (a == 52) and (h == 100):
        a1 = qp132
    elif (a == 53) and (h == 100):
        a1 = qp133
    elif (a == 54) and (h == 100):
        a1 = qp134
    elif (a == 55) and (h == 100):
        a1 = qp135
    elif (a == 56) and (h == 100):
        a1 = qp136
    elif (a == 57) and (h == 100):
        a1 = qp137
    elif (a == 58) and (h == 100):
        a1 = qp138
    elif (a == 59) and (h == 100):
        a1 = qp139
    elif (a == 60) and (h == 100):
        a1 = qp140
    elif (a == 61) and (h == 100):
        a1 = qp141
    elif (a == 62) and (h == 100):
        a1 = qp142
    elif (a == 63) and (h == 100):
        a1 = qp143
    elif (a == 64) and (h == 100):
        a1 = qp144
    elif (a == 65) and (h == 100):
        a1 = qp145
    elif (a == 66) and (h == 100):
        a1 = qp146
    elif (a == 67) and (h == 100):
        a1 = qp147
    elif (a == 68) and (h == 100):
        a1 = qp148
    elif (a == 69) and (h == 100):
        a1 = qp149
    elif (a == 70) and (h == 100):
        a1 = qp150
    elif (a == 71) and (h == 100):
        a1 = qp151
    elif (a == 72) and (h == 100):
        a1 = qp152
    elif (a == 73) and (h == 100):
        a1 = qp153
    elif (a == 74) and (h == 100):
        a1 = qp154
    elif (a == 75) and (h == 100):
        a1 = qp155
    elif (a == 76) and (h == 100):
        a1 = qp156
    elif (a == 77) and (h == 100):
        a1 = qp157
    elif (a == 78) and (h == 100):
        a1 = qp158
    elif (a == 79) and (h == 100):
        a1 = qp159
    elif (a >= 80) and (a <= 100) and (h == 100):
        a1 = qp160
    else:
        a1 = qp20
    return a1


if LMS.last_semester == 1:
    # ---- Semester 1----- #

    total_num_semester_1 = []
    total_ch_semester_1 = []
    for row in range(5, 5 + LMS.number_of_subjects_1):
        total = int(ws[f"K{row}"].value)
        ch_raw = int(ws[f"F{row}"].value)
        ch_str = ch_raw[0]
        ch = int(ch_str)
        total_ch_semester_1.append(ch)
        total_num_semester_1.append(total)
    total_ch_sum = sum(total_ch_semester_1)

    def tm1():
        total_mark_semester_1 = []
        for items in total_ch_semester_1:
            if items == 1:
                h = 20
            elif items == 2:
                h = 40
            elif items == 3:
                h = 60
            elif items == 4:
                h = 80
            elif items == 5:
                h = 100
            total_mark_semester_1.append(h)
        return total_mark_semester_1

    tm1()
    total_qp_semester_1 = []
    for (a, h) in zip(total_num_semester_1, tm1()):
        doit()
        total_qp_semester_1.append(doit())
    total_qp_semester_1_sum = sum(total_qp_semester_1)
    gpa = round(total_qp_semester_1_sum / total_ch_sum, 2)

    ws[f"E{last_row + 1}"] = "Semester 1 GPA:"
    ws[f"E{last_row + 1}"].font = Font(bold=True, size=11, name="Google Sans")
    ws.cell(row=last_row + 1, column=2).border = LMS.thin_border
    ws[f"F{last_row + 1}"] = gpa
    ws[f"F{last_row + 1}"].font = Font(bold=True, size=11, name="Google Sans")
    ws[f"F{last_row + 1}"].alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=last_row + 1, column=3).border = LMS.thin_border

    total_sum_1 = sum(total_num_semester_1)
    ws[f"B{last_row + 1}"] = "Semester 1 marks:"
    ws[f"B{last_row + 1}"].font = Font(bold=True, size=11, name="Google Sans")
    ws.cell(row=last_row + 1, column=2).border = LMS.thin_border
    ws[f"C{last_row + 1}"] = total_sum_1
    ws[f"C{last_row + 1}"].font = Font(bold=True, size=11, name="Google Sans")
    ws[f"C{last_row + 1}"].alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=last_row + 1, column=3).border = LMS.thin_border

    # ----- Semester 1 ------#
elif LMS.last_semester == 2:
    # ---- Semester 1----- #

    total_num_semester_1 = []
    total_ch_semester_1 = []
    for row in range(5, 5 + LMS.number_of_subjects_1):
        total = int(ws[f"K{row}"].value)
        ch_raw = str(ws[f"F{row}"].value)
        ch_str = ch_raw[0]
        ch = int(ch_str)
        total_ch_semester_1.append(ch)
        total_num_semester_1.append(total)
    total_ch_sum = sum(total_ch_semester_1)


    def tm1():
        total_mark_semester_1 = []
        for items in total_ch_semester_1:
            if items == 1:
                h = 20
            elif items == 2:
                h = 40
            elif items == 3:
                h = 60
            elif items == 4:
                h = 80
            elif items == 5:
                h = 100
            total_mark_semester_1.append(h)
        return total_mark_semester_1

    tm1()
    total_qp_semester_1 = []
    for (a, h) in zip(total_num_semester_1, tm1()):
        doit()
        total_qp_semester_1.append(doit())
    total_qp_semester_1_sum = sum(total_qp_semester_1)
    gpa = round(total_qp_semester_1_sum / total_ch_sum, 2)




    ws[f"E{last_row + 1}"] = "Semester 1 GPA:"
    ws[f"E{last_row + 1}"].font = Font(bold=True, size=11, name="Google Sans")
    # ws.cell(row=last_row + 1, column=5).border = LMS.thin_border
    ws[f"F{last_row + 1}"] = gpa
    ws[f"F{last_row + 1}"].font = Font(bold=True, size=11, name="Google Sans")
    ws[f"F{last_row + 1}"].alignment = Alignment(horizontal='center', vertical='center')
    # ws.cell(row=last_row + 1, column=6).border = LMS.thin_border

    total_sum_1 = sum(total_num_semester_1)
    ws[f"B{last_row + 1}"] = "Semester 1 marks:"
    ws[f"B{last_row + 1}"].font = Font(bold=True, size=11, name="Google Sans")
    ws.cell(row=last_row + 1, column=2).border = LMS.thin_border
    ws[f"C{last_row + 1}"] = total_sum_1
    ws[f"C{last_row + 1}"].font = Font(bold=True, size=11, name="Google Sans")
    ws[f"C{last_row + 1}"].alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=last_row + 1, column=3).border = LMS.thin_border

    # ----- Semester 1 ------#

    # ----- Semester 2 ----- #
    total_num_semester_2 = []
    total_ch_semester_2 = []
    first = 5 + LMS.number_of_subjects_1
    second = first + LMS.number_of_subjects_2
    for row in range(first, second):
        total2 = int(ws[f"K{row}"].value)
        ch_raw2 = str(ws[f"F{row}"].value)
        ch_str2 = ch_raw2[0]
        ch2 = int(ch_str2)
        total_ch_semester_2.append(ch2)
        total_num_semester_2.append(total2)
    total_ch_sum2 = sum(total_ch_semester_2)

    def tm2():
        total_mark_semester_2 = []
        for items in total_ch_semester_2:
            if items == 1:
                h = 20
            elif items == 2:
                h = 40
            elif items == 3:
                h = 60
            elif items == 4:
                h = 80
            elif items == 5:
                h = 100
            total_mark_semester_2.append(h)
        return total_mark_semester_2


    tm2()
    total_qp_semester_2 = []
    for (a, h) in zip(total_num_semester_2, tm2()):
        doit()
        total_qp_semester_2.append(doit())
    total_qp_semester_2_sum = sum(total_qp_semester_2)
    gpa2 = round(total_qp_semester_2_sum / total_ch_sum2, 2)


    ws[f"E{last_row + 2}"] = "Semester 2 GPA:"
    ws[f"E{last_row + 2}"].font = Font(bold=True, size=11, name="Google Sans")
    # ws.cell(row=last_row + 2, column=5).border = LMS.thin_border
    ws[f"F{last_row + 2}"] = gpa2
    ws[f"F{last_row + 2}"].font = Font(bold=True, size=11, name="Google Sans")
    ws[f"F{last_row + 2}"].alignment = Alignment(horizontal='center', vertical='center')
    # ws.cell(row=last_row + 2, column=6).border = LMS.thin_border

    total_sum_2 = sum(total_num_semester_2)
    ws[f"B{last_row + 2}"] = "Semester 2 marks:"
    ws[f"B{last_row + 2}"].font = Font(bold=True, size=11, name="Google Sans")
    ws.cell(row=last_row + 2, column=2).border = LMS.thin_border
    ws[f"C{last_row + 2}"] = total_sum_2
    ws[f"C{last_row + 2}"].font = Font(bold=True, size=11, name="Google Sans")
    ws[f"C{last_row + 2}"].alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=last_row + 2, column=3).border = LMS.thin_border
    # ----- Semester 2 -----#


elif LMS.last_semester == 3:
    # ---- Semester 1----- #

    total_num_semester_1 = []
    total_ch_semester_1 = []
    for row in range(5, 5 + LMS.number_of_subjects_1):
        total = int(ws[f"K{row}"].value)
        ch_raw = str(ws[f"F{row}"].value)
        ch_str = ch_raw[0]
        ch = int(ch_str)
        total_ch_semester_1.append(ch)
        total_num_semester_1.append(total)
    total_ch_sum = sum(total_ch_semester_1)


    def tm():
        total_mark_semester_1 = []
        for items in total_ch_semester_1:
            if items == 1:
                h = 20
            elif items == 2:
                h = 40
            elif items == 3:
                h = 60
            elif items == 4:
                h = 80
            elif items == 5:
                h = 100
            total_mark_semester_1.append(h)
        return total_mark_semester_1


    tm()
    total_qp_semester_1 = []
    for (a, h) in zip(total_num_semester_1, tm()):
        doit()
        total_qp_semester_1.append(doit())
    total_qp_semester_1_sum = sum(total_qp_semester_1)
    gpa = round(total_qp_semester_1_sum / total_ch_sum, 2)

    ws[f"E{last_row + 1}"] = "Semester 1 GPA:"
    ws[f"E{last_row + 1}"].font = Font(bold=True, size=11, name="Google Sans")
    # ws.cell(row=last_row + 1, column=5).border = LMS.thin_border
    ws[f"F{last_row + 1}"] = gpa
    ws[f"F{last_row + 1}"].font = Font(bold=True, size=11, name="Google Sans")
    ws[f"F{last_row + 1}"].alignment = Alignment(horizontal='center', vertical='center')
    # ws.cell(row=last_row + 1, column=6).border = LMS.thin_border

    total_sum_1 = sum(total_num_semester_1)
    ws[f"B{last_row + 1}"] = "Semester 1 marks:"
    ws[f"B{last_row + 1}"].font = Font(bold=True, size=11, name="Google Sans")
    ws.cell(row=last_row + 1, column=2).border = LMS.thin_border
    ws[f"C{last_row + 1}"] = total_sum_1
    ws[f"C{last_row + 1}"].font = Font(bold=True, size=11, name="Google Sans")
    ws[f"C{last_row + 1}"].alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=last_row + 1, column=3).border = LMS.thin_border

    # ----- Semester 1 ------#

    # ----- Semester 2 ----- #
    total_num_semester_2 = []
    total_ch_semester_2 = []
    first = 5 + LMS.number_of_subjects_1
    second = first + LMS.number_of_subjects_2
    for row in range(first, second):
        total2 = int(ws[f"K{row}"].value)
        ch_raw2 = str(ws[f"F{row}"].value)
        ch_str2 = ch_raw2[0]
        ch2 = int(ch_str2)
        total_ch_semester_2.append(ch2)
        total_num_semester_2.append(total2)
    total_ch_sum2 = sum(total_ch_semester_2)


    def tm2():
        total_mark_semester_2 = []
        for items in total_ch_semester_2:
            if items == 1:
                h = 20
            elif items == 2:
                h = 40
            elif items == 3:
                h = 60
            elif items == 4:
                h = 80
            elif items == 5:
                h = 100
            total_mark_semester_2.append(h)
        return total_mark_semester_2


    tm2()
    total_qp_semester_2 = []
    for (a, h) in zip(total_num_semester_2, tm2()):
        doit()
        total_qp_semester_2.append(doit())
    total_qp_semester_2_sum = sum(total_qp_semester_2)
    gpa2 = round(total_qp_semester_2_sum / total_ch_sum2, 2)

    ws[f"E{last_row + 2}"] = "Semester 2 GPA:"
    ws[f"E{last_row + 2}"].font = Font(bold=True, size=11, name="Google Sans")
    # ws.cell(row=last_row + 2, column=5).border = LMS.thin_border
    ws[f"F{last_row + 2}"] = gpa2
    ws[f"F{last_row + 2}"].font = Font(bold=True, size=11, name="Google Sans")
    ws[f"F{last_row + 2}"].alignment = Alignment(horizontal='center', vertical='center')
    # ws.cell(row=last_row + 2, column=6).border = LMS.thin_border

    total_sum_2 = sum(total_num_semester_2)
    ws[f"B{last_row + 2}"] = "Semester 2 marks:"
    ws[f"B{last_row + 2}"].font = Font(bold=True, size=11, name="Google Sans")
    ws.cell(row=last_row + 2, column=2).border = LMS.thin_border
    ws[f"C{last_row + 2}"] = total_sum_2
    ws[f"C{last_row + 2}"].font = Font(bold=True, size=11, name="Google Sans")
    ws[f"C{last_row + 2}"].alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=last_row + 2, column=3).border = LMS.thin_border
    # ----- Semester 2 -----#

    # ----- Semester 3 ----- #
    total_num_semester_3 = []
    total_ch_semester_3 = []
    first = 5 + LMS.number_of_subjects_1
    second = first + LMS.number_of_subjects_2
    third = second + LMS.number_of_subjects_3
    for row in range(second, third):
        total3 = int(ws[f"K{row}"].value)
        ch_raw3 = str(ws[f"F{row}"].value)
        ch_str3 = ch_raw3[0]
        ch3 = int(ch_str3)
        total_ch_semester_3.append(ch3)
        total_num_semester_3.append(total3)
    total_ch_sum3 = sum(total_ch_semester_3)

    def tm3():
        total_mark_semester_3 = []
        for items in total_ch_semester_3:
            if items == 1:
                h = 20
            elif items == 2:
                h = 40
            elif items == 3:
                h = 60
            elif items == 4:
                h = 80
            elif items == 5:
                h = 100
            total_mark_semester_3.append(h)
            return total_mark_semester_3


    tm3()
    total_qp_semester_3 = []
    for (a, h) in zip(total_num_semester_3, tm3()):
        doit()
        total_qp_semester_3.append(doit())
    total_qp_semester_3_sum = sum(total_qp_semester_3)
    gpa3 = round(total_qp_semester_3_sum / total_ch_sum3, 2)

    ws[f"E{last_row + 3}"] = "Semester 3 GPA:"
    ws[f"E{last_row + 3}"].font = Font(bold=True, size=11, name="Google Sans")
    # ws.cell(row=last_row + 2, column=5).border = LMS.thin_border
    ws[f"F{last_row + 3}"] = gpa3
    ws[f"F{last_row + 3}"].font = Font(bold=True, size=11, name="Google Sans")
    ws[f"F{last_row + 3}"].alignment = Alignment(horizontal='center', vertical='center')
    # ws.cell(row=last_row + 2, column=6).border = LMS.thin_border

    total_sum_3 = sum(total_num_semester_3)
    ws[f"B{last_row + 3}"] = "Semester 3 marks:"
    ws[f"B{last_row + 3}"].font = Font(bold=True, size=11, name="Google Sans")
    ws.cell(row=last_row + 3, column=2).border = LMS.thin_border
    ws[f"C{last_row + 3}"] = total_sum_3
    ws[f"C{last_row + 3}"].font = Font(bold=True, size=11, name="Google Sans")
    ws[f"C{last_row + 3}"].alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=last_row + 3, column=3).border = LMS.thin_border
    # ----- Semester 3 ----- #

elif LMS.last_semester == 4:
    # ---- Semester 1----- #
    total_num_semester_1 = []
    for row in range(5, 5 + LMS.number_of_subjects_1):
        total = int(ws[f"K{row}"].value)
        total_num_semester_1.append(total)
    total_sum_1 = sum(total_num_semester_1)
    ws[f"B{last_row + 1}"] = "Semester 1 marks:"
    ws[f"B{last_row + 1}"].font = Font(bold=True, size=11, name="Google Sans")
    ws.cell(row=last_row + 1, column=2).border = LMS.thin_border
    ws[f"C{last_row + 1}"] = total_sum_1
    ws[f"C{last_row + 1}"].font = Font(bold=True, size=11, name="Google Sans")
    ws[f"C{last_row + 1}"].alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=last_row + 1, column=3).border = LMS.thin_border

    # ----- Semester 1 ------#

    # ----- Semester 2 ----- #
    total_num_semester_2 = []
    first = 5 + LMS.number_of_subjects_1
    second = first + LMS.number_of_subjects_2
    for row in range(first, second):
        total2 = int(ws[f"K{row}"].value)
        total_num_semester_2.append(total2)
    total_sum_2 = sum(total_num_semester_2)
    ws[f"B{last_row + 2}"] = "Semester 2 marks:"
    ws[f"B{last_row + 2}"].font = Font(bold=True, size=11, name="Google Sans")
    ws.cell(row=last_row + 2, column=2).border = LMS.thin_border
    ws[f"C{last_row + 2}"] = total_sum_2
    ws[f"C{last_row + 2}"].font = Font(bold=True, size=11, name="Google Sans")
    ws[f"C{last_row + 2}"].alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=last_row + 2, column=3).border = LMS.thin_border
    # ----- Semester 2 ----- #

    # ----- Semester 3 ----- #
    total_num_semester_3 = []
    first = 5 + LMS.number_of_subjects_1
    second = first + LMS.number_of_subjects_2
    third = second + LMS.number_of_subjects_3
    for row in range(second, third):
        total3 = int(ws[f"K{row}"].value)
        total_num_semester_3.append(total3)
    total_sum_3 = sum(total_num_semester_3)
    ws[f"B{last_row + 3}"] = "Semester 3 marks:"
    ws[f"B{last_row + 3}"].font = Font(bold=True, size=11, name="Google Sans")
    ws.cell(row=last_row + 3, column=2).border = LMS.thin_border
    ws[f"C{last_row + 3}"] = total_sum_3
    ws[f"C{last_row + 3}"].font = Font(bold=True, size=11, name="Google Sans")
    ws[f"C{last_row + 3}"].alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=last_row + 3, column=3).border = LMS.thin_border
    # ----- Semester 3 ----- #

    # ----- Semester 4 ----- #
    total_num_semester_4 = []
    first = 5 + LMS.number_of_subjects_1
    second = first + LMS.number_of_subjects_2
    third = second + LMS.number_of_subjects_3
    fourth = third + LMS.number_of_subjects_4
    for row in range(third, fourth):
        total4 = int(ws[f"K{row}"].value)
        total_num_semester_4.append(total4)
    total_sum_4 = sum(total_num_semester_4)
    ws[f"B{last_row + 4}"] = "Semester 4 marks:"
    ws[f"B{last_row + 4}"].font = Font(bold=True, size=11, name="Google Sans")
    ws.cell(row=last_row + 4, column=2).border = LMS.thin_border
    ws[f"C{last_row + 4}"] = total_sum_4
    ws[f"C{last_row + 4}"].font = Font(bold=True, size=11, name="Google Sans")
    ws[f"C{last_row + 4}"].alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=last_row + 4, column=3).border = LMS.thin_border
    # ----- Semester 4 ----- #

elif LMS.last_semester == 5:
    # ---- Semester 1----- #
    total_num_semester_1 = []
    for row in range(5, 5 + LMS.number_of_subjects_1):
        total = int(ws[f"K{row}"].value)
        total_num_semester_1.append(total)
    total_sum_1 = sum(total_num_semester_1)
    ws[f"B{last_row + 1}"] = "Semester 1 marks:"
    ws[f"B{last_row + 1}"].font = Font(bold=True, size=11, name="Google Sans")
    ws.cell(row=last_row + 1, column=2).border = LMS.thin_border
    ws[f"C{last_row + 1}"] = total_sum_1
    ws[f"C{last_row + 1}"].font = Font(bold=True, size=11, name="Google Sans")
    ws[f"C{last_row + 1}"].alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=last_row + 1, column=3).border = LMS.thin_border

    # ----- Semester 1 ------#

    # ----- Semester 2 ----- #
    total_num_semester_2 = []
    first = 5 + LMS.number_of_subjects_1
    second = first + LMS.number_of_subjects_2
    for row in range(first, second):
        total2 = int(ws[f"K{row}"].value)
        total_num_semester_2.append(total2)
    total_sum_2 = sum(total_num_semester_2)
    ws[f"B{last_row + 2}"] = "Semester 2 marks:"
    ws[f"B{last_row + 2}"].font = Font(bold=True, size=11, name="Google Sans")
    ws.cell(row=last_row + 2, column=2).border = LMS.thin_border
    ws[f"C{last_row + 2}"] = total_sum_2
    ws[f"C{last_row + 2}"].font = Font(bold=True, size=11, name="Google Sans")
    ws[f"C{last_row + 2}"].alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=last_row + 2, column=3).border = LMS.thin_border
    # ----- Semester 2 ----- #

    # ----- Semester 3 ----- #
    total_num_semester_3 = []
    first = 5 + LMS.number_of_subjects_1
    second = first + LMS.number_of_subjects_2
    third = second + LMS.number_of_subjects_3
    for row in range(second, third):
        total3 = int(ws[f"K{row}"].value)
        total_num_semester_3.append(total3)
    total_sum_3 = sum(total_num_semester_3)
    ws[f"B{last_row + 3}"] = "Semester 3 marks:"
    ws[f"B{last_row + 3}"].font = Font(bold=True, size=11, name="Google Sans")
    ws.cell(row=last_row + 3, column=2).border = LMS.thin_border
    ws[f"C{last_row + 3}"] = total_sum_3
    ws[f"C{last_row + 3}"].font = Font(bold=True, size=11, name="Google Sans")
    ws[f"C{last_row + 3}"].alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=last_row + 3, column=3).border = LMS.thin_border
    # ----- Semester 3 ----- #

    # ----- Semester 4 ----- #
    total_num_semester_4 = []
    first = 5 + LMS.number_of_subjects_1
    second = first + LMS.number_of_subjects_2
    third = second + LMS.number_of_subjects_3
    fourth = third + LMS.number_of_subjects_4
    for row in range(third, fourth):
        total4 = int(ws[f"K{row}"].value)
        total_num_semester_4.append(total4)
    total_sum_4 = sum(total_num_semester_4)
    ws[f"B{last_row + 4}"] = "Semester 4 marks:"
    ws[f"B{last_row + 4}"].font = Font(bold=True, size=11, name="Google Sans")
    ws.cell(row=last_row + 4, column=2).border = LMS.thin_border
    ws[f"C{last_row + 4}"] = total_sum_4
    ws[f"C{last_row + 4}"].font = Font(bold=True, size=11, name="Google Sans")
    ws[f"C{last_row + 4}"].alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=last_row + 4, column=3).border = LMS.thin_border
    # ----- Semester 4 ----- #

    # ----- Semester 5 ----- #
    total_num_semester_5 = []
    first = 5 + LMS.number_of_subjects_1
    second = first + LMS.number_of_subjects_2
    third = second + LMS.number_of_subjects_3
    fourth = third + LMS.number_of_subjects_4
    fifth = fourth + LMS.number_of_subjects_5
    for row in range(fourth, fifth):
        total5 = int(ws[f"K{row}"].value)
        total_num_semester_5.append(total5)
    total_sum_5 = sum(total_num_semester_5)
    ws[f"B{last_row + 5}"] = "Semester 5 marks:"
    ws[f"B{last_row + 5}"].font = Font(bold=True, size=11, name="Google Sans")
    ws.cell(row=last_row + 5, column=2).border = LMS.thin_border
    ws[f"C{last_row + 5}"] = total_sum_5
    ws[f"C{last_row + 5}"].font = Font(bold=True, size=11, name="Google Sans")
    ws[f"C{last_row + 5}"].alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=last_row + 5, column=3).border = LMS.thin_border
    # ----- Semester 5 ----- #

elif LMS.last_semester == 6:
    # ---- Semester 1----- #
    total_num_semester_1 = []
    for row in range(5, 5 + LMS.number_of_subjects_1):
        total = int(ws[f"K{row}"].value)
        total_num_semester_1.append(total)
    total_sum_1 = sum(total_num_semester_1)
    ws[f"B{last_row + 1}"] = "Semester 1 marks:"
    ws[f"B{last_row + 1}"].font = Font(bold=True, size=11, name="Google Sans")
    ws.cell(row=last_row + 1, column=2).border = LMS.thin_border
    ws[f"C{last_row + 1}"] = total_sum_1
    ws[f"C{last_row + 1}"].font = Font(bold=True, size=11, name="Google Sans")
    ws[f"C{last_row + 1}"].alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=last_row + 1, column=3).border = LMS.thin_border

    # ----- Semester 1 ------#

    # ----- Semester 2 ----- #
    total_num_semester_2 = []
    first = 5 + LMS.number_of_subjects_1
    second = first + LMS.number_of_subjects_2
    for row in range(first, second):
        total2 = int(ws[f"K{row}"].value)
        total_num_semester_2.append(total2)
    total_sum_2 = sum(total_num_semester_2)
    ws[f"B{last_row + 2}"] = "Semester 2 marks:"
    ws[f"B{last_row + 2}"].font = Font(bold=True, size=11, name="Google Sans")
    ws.cell(row=last_row + 2, column=2).border = LMS.thin_border
    ws[f"C{last_row + 2}"] = total_sum_2
    ws[f"C{last_row + 2}"].font = Font(bold=True, size=11, name="Google Sans")
    ws[f"C{last_row + 2}"].alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=last_row + 2, column=3).border = LMS.thin_border
    # ----- Semester 2 ----- #

    # ----- Semester 3 ----- #
    total_num_semester_3 = []
    first = 5 + LMS.number_of_subjects_1
    second = first + LMS.number_of_subjects_2
    third = second + LMS.number_of_subjects_3
    for row in range(second, third):
        total3 = int(ws[f"K{row}"].value)
        total_num_semester_3.append(total3)
    total_sum_3 = sum(total_num_semester_3)
    ws[f"B{last_row + 3}"] = "Semester 3 marks:"
    ws[f"B{last_row + 3}"].font = Font(bold=True, size=11, name="Google Sans")
    ws.cell(row=last_row + 3, column=2).border = LMS.thin_border
    ws[f"C{last_row + 3}"] = total_sum_3
    ws[f"C{last_row + 3}"].font = Font(bold=True, size=11, name="Google Sans")
    ws[f"C{last_row + 3}"].alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=last_row + 3, column=3).border = LMS.thin_border
    # ----- Semester 3 ----- #

    # ----- Semester 4 ----- #
    total_num_semester_4 = []
    first = 5 + LMS.number_of_subjects_1
    second = first + LMS.number_of_subjects_2
    third = second + LMS.number_of_subjects_3
    fourth = third + LMS.number_of_subjects_4
    for row in range(third, fourth):
        total4 = int(ws[f"K{row}"].value)
        total_num_semester_4.append(total4)
    total_sum_4 = sum(total_num_semester_4)
    ws[f"B{last_row + 4}"] = "Semester 4 marks:"
    ws[f"B{last_row + 4}"].font = Font(bold=True, size=11, name="Google Sans")
    ws.cell(row=last_row + 4, column=2).border = LMS.thin_border
    ws[f"C{last_row + 4}"] = total_sum_4
    ws[f"C{last_row + 4}"].font = Font(bold=True, size=11, name="Google Sans")
    ws[f"C{last_row + 4}"].alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=last_row + 4, column=3).border = LMS.thin_border
    # ----- Semester 4 ----- #

    # ----- Semester 5 ----- #
    total_num_semester_5 = []
    first = 5 + LMS.number_of_subjects_1
    second = first + LMS.number_of_subjects_2
    third = second + LMS.number_of_subjects_3
    fourth = third + LMS.number_of_subjects_4
    fifth = fourth + LMS.number_of_subjects_5
    for row in range(fourth, fifth):
        total5 = int(ws[f"K{row}"].value)
        total_num_semester_5.append(total5)
    total_sum_5 = sum(total_num_semester_5)
    ws[f"B{last_row + 5}"] = "Semester 5 marks:"
    ws[f"B{last_row + 5}"].font = Font(bold=True, size=11, name="Google Sans")
    ws.cell(row=last_row + 5, column=2).border = LMS.thin_border
    ws[f"C{last_row + 5}"] = total_sum_5
    ws[f"C{last_row + 5}"].font = Font(bold=True, size=11, name="Google Sans")
    ws[f"C{last_row + 5}"].alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=last_row + 5, column=3).border = LMS.thin_border
    # ----- Semester 5 ----- #

    # ----- Semester 6 ----- #
    total_num_semester_6 = []
    first = 5 + LMS.number_of_subjects_1
    second = first + LMS.number_of_subjects_2
    third = second + LMS.number_of_subjects_3
    fourth = third + LMS.number_of_subjects_4
    fifth = fourth + LMS.number_of_subjects_5
    sixth = fifth + LMS.number_of_subjects_6
    for row in range(fifth, sixth):
        total6 = int(ws[f"K{row}"].value)
        total_num_semester_6.append(total6)
    total_sum_6 = sum(total_num_semester_6)
    ws[f"B{last_row + 6}"] = "Semester 6 marks:"
    ws[f"B{last_row + 6}"].font = Font(bold=True, size=11, name="Google Sans")
    ws.cell(row=last_row + 6, column=2).border = LMS.thin_border
    ws[f"C{last_row + 6}"] = total_sum_6
    ws[f"C{last_row + 6}"].font = Font(bold=True, size=11, name="Google Sans")
    ws[f"C{last_row + 6}"].alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=last_row + 6, column=3).border = LMS.thin_border
    # ----- Semester 6 ---- #

elif LMS.last_semester == 7:
    # ---- Semester 1----- #
    total_num_semester_1 = []
    for row in range(5, 5 + LMS.number_of_subjects_1):
        total = int(ws[f"K{row}"].value)
        total_num_semester_1.append(total)
    total_sum_1 = sum(total_num_semester_1)
    ws[f"B{last_row + 1}"] = "Semester 1 marks:"
    ws[f"B{last_row + 1}"].font = Font(bold=True, size=11, name="Google Sans")
    ws.cell(row=last_row + 1, column=2).border = LMS.thin_border
    ws[f"C{last_row + 1}"] = total_sum_1
    ws[f"C{last_row + 1}"].font = Font(bold=True, size=11, name="Google Sans")
    ws[f"C{last_row + 1}"].alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=last_row + 1, column=3).border = LMS.thin_border

    # ----- Semester 1 ------#

    # ----- Semester 2 ----- #
    total_num_semester_2 = []
    first = 5 + LMS.number_of_subjects_1
    second = first + LMS.number_of_subjects_2
    for row in range(first, second):
        total2 = int(ws[f"K{row}"].value)
        total_num_semester_2.append(total2)
    total_sum_2 = sum(total_num_semester_2)
    ws[f"B{last_row + 2}"] = "Semester 2 marks:"
    ws[f"B{last_row + 2}"].font = Font(bold=True, size=11, name="Google Sans")
    ws.cell(row=last_row + 2, column=2).border = LMS.thin_border
    ws[f"C{last_row + 2}"] = total_sum_2
    ws[f"C{last_row + 2}"].font = Font(bold=True, size=11, name="Google Sans")
    ws[f"C{last_row + 2}"].alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=last_row + 2, column=3).border = LMS.thin_border
    # ----- Semester 2 ----- #

    # ----- Semester 3 ----- #
    total_num_semester_3 = []
    first = 5 + LMS.number_of_subjects_1
    second = first + LMS.number_of_subjects_2
    third = second + LMS.number_of_subjects_3
    for row in range(second, third):
        total3 = int(ws[f"K{row}"].value)
        total_num_semester_3.append(total3)
    total_sum_3 = sum(total_num_semester_3)
    ws[f"B{last_row + 3}"] = "Semester 3 marks:"
    ws[f"B{last_row + 3}"].font = Font(bold=True, size=11, name="Google Sans")
    ws.cell(row=last_row + 3, column=2).border = LMS.thin_border
    ws[f"C{last_row + 3}"] = total_sum_3
    ws[f"C{last_row + 3}"].font = Font(bold=True, size=11, name="Google Sans")
    ws[f"C{last_row + 3}"].alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=last_row + 3, column=3).border = LMS.thin_border
    # ----- Semester 3 ----- #

    # ----- Semester 4 ----- #
    total_num_semester_4 = []
    first = 5 + LMS.number_of_subjects_1
    second = first + LMS.number_of_subjects_2
    third = second + LMS.number_of_subjects_3
    fourth = third + LMS.number_of_subjects_4
    for row in range(third, fourth):
        total4 = int(ws[f"K{row}"].value)
        total_num_semester_4.append(total4)
    total_sum_4 = sum(total_num_semester_4)
    ws[f"B{last_row + 4}"] = "Semester 4 marks:"
    ws[f"B{last_row + 4}"].font = Font(bold=True, size=11, name="Google Sans")
    ws.cell(row=last_row + 4, column=2).border = LMS.thin_border
    ws[f"C{last_row + 4}"] = total_sum_4
    ws[f"C{last_row + 4}"].font = Font(bold=True, size=11, name="Google Sans")
    ws[f"C{last_row + 4}"].alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=last_row + 4, column=3).border = LMS.thin_border
    # ----- Semester 4 ----- #

    # ----- Semester 5 ----- #
    total_num_semester_5 = []
    first = 5 + LMS.number_of_subjects_1
    second = first + LMS.number_of_subjects_2
    third = second + LMS.number_of_subjects_3
    fourth = third + LMS.number_of_subjects_4
    fifth = fourth + LMS.number_of_subjects_5
    for row in range(fourth, fifth):
        total5 = int(ws[f"K{row}"].value)
        total_num_semester_5.append(total5)
    total_sum_5 = sum(total_num_semester_5)
    ws[f"B{last_row + 5}"] = "Semester 5 marks:"
    ws[f"B{last_row + 5}"].font = Font(bold=True, size=11, name="Google Sans")
    ws.cell(row=last_row + 5, column=2).border = LMS.thin_border
    ws[f"C{last_row + 5}"] = total_sum_5
    ws[f"C{last_row + 5}"].font = Font(bold=True, size=11, name="Google Sans")
    ws[f"C{last_row + 5}"].alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=last_row + 5, column=3).border = LMS.thin_border
    # ----- Semester 5 ----- #

    # ----- Semester 6 ----- #
    total_num_semester_6 = []
    first = 5 + LMS.number_of_subjects_1
    second = first + LMS.number_of_subjects_2
    third = second + LMS.number_of_subjects_3
    fourth = third + LMS.number_of_subjects_4
    fifth = fourth + LMS.number_of_subjects_5
    sixth = fifth + LMS.number_of_subjects_6
    for row in range(fifth, sixth):
        total6 = int(ws[f"K{row}"].value)
        total_num_semester_6.append(total6)
    total_sum_6 = sum(total_num_semester_6)
    ws[f"B{last_row + 6}"] = "Semester 6 marks:"
    ws[f"B{last_row + 6}"].font = Font(bold=True, size=11, name="Google Sans")
    ws.cell(row=last_row + 6, column=2).border = LMS.thin_border
    ws[f"C{last_row + 6}"] = total_sum_6
    ws[f"C{last_row + 6}"].font = Font(bold=True, size=11, name="Google Sans")
    ws[f"C{last_row + 6}"].alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=last_row + 6, column=3).border = LMS.thin_border
    # ----- Semester 6 ---- #

    # ----- Semester 7 ----- #
    total_num_semester_7 = []
    first = 5 + LMS.number_of_subjects_1
    second = first + LMS.number_of_subjects_2
    third = second + LMS.number_of_subjects_3
    fourth = third + LMS.number_of_subjects_4
    fifth = fourth + LMS.number_of_subjects_5
    sixth = fifth + LMS.number_of_subjects_6
    seventh = sixth + LMS.number_of_subjects_7
    for row in range(sixth, seventh):
        total7 = int(ws[f"K{row}"].value)
        total_num_semester_7.append(total7)
    total_sum_7 = sum(total_num_semester_7)
    ws[f"B{last_row + 7}"] = "Semester 7 marks:"
    ws[f"B{last_row + 7}"].font = Font(bold=True, size=11, name="Google Sans")
    ws.cell(row=last_row + 7, column=2).border = LMS.thin_border
    ws[f"C{last_row + 7}"] = total_sum_7
    ws[f"C{last_row + 7}"].font = Font(bold=True, size=11, name="Google Sans")
    ws[f"C{last_row + 7}"].alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=last_row + 7, column=3).border = LMS.thin_border
    # ----- Semester 7 ---- #

elif LMS.last_semester == 7:
    # ---- Semester 1----- #
    total_num_semester_1 = []
    for row in range(5, 5 + LMS.number_of_subjects_1):
        total = int(ws[f"K{row}"].value)
        total_num_semester_1.append(total)
    total_sum_1 = sum(total_num_semester_1)
    ws[f"B{last_row + 1}"] = "Semester 1 marks:"
    ws[f"B{last_row + 1}"].font = Font(bold=True, size=11, name="Google Sans")
    ws.cell(row=last_row + 1, column=2).border = LMS.thin_border
    ws[f"C{last_row + 1}"] = total_sum_1
    ws[f"C{last_row + 1}"].font = Font(bold=True, size=11, name="Google Sans")
    ws[f"C{last_row + 1}"].alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=last_row + 1, column=3).border = LMS.thin_border

    # ----- Semester 1 ------#

    # ----- Semester 2 ----- #
    total_num_semester_2 = []
    first = 5 + LMS.number_of_subjects_1
    second = first + LMS.number_of_subjects_2
    for row in range(first, second):
        total2 = int(ws[f"K{row}"].value)
        total_num_semester_2.append(total2)
    total_sum_2 = sum(total_num_semester_2)
    ws[f"B{last_row + 2}"] = "Semester 2 marks:"
    ws[f"B{last_row + 2}"].font = Font(bold=True, size=11, name="Google Sans")
    ws.cell(row=last_row + 2, column=2).border = LMS.thin_border
    ws[f"C{last_row + 2}"] = total_sum_2
    ws[f"C{last_row + 2}"].font = Font(bold=True, size=11, name="Google Sans")
    ws[f"C{last_row + 2}"].alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=last_row + 2, column=3).border = LMS.thin_border
    # ----- Semester 2 ----- #

    # ----- Semester 3 ----- #
    total_num_semester_3 = []
    first = 5 + LMS.number_of_subjects_1
    second = first + LMS.number_of_subjects_2
    third = second + LMS.number_of_subjects_3
    for row in range(second, third):
        total3 = int(ws[f"K{row}"].value)
        total_num_semester_3.append(total3)
    total_sum_3 = sum(total_num_semester_3)
    ws[f"B{last_row + 3}"] = "Semester 3 marks:"
    ws[f"B{last_row + 3}"].font = Font(bold=True, size=11, name="Google Sans")
    ws.cell(row=last_row + 3, column=2).border = LMS.thin_border
    ws[f"C{last_row + 3}"] = total_sum_3
    ws[f"C{last_row + 3}"].font = Font(bold=True, size=11, name="Google Sans")
    ws[f"C{last_row + 3}"].alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=last_row + 3, column=3).border = LMS.thin_border
    # ----- Semester 3 ----- #

    # ----- Semester 4 ----- #
    total_num_semester_4 = []
    first = 5 + LMS.number_of_subjects_1
    second = first + LMS.number_of_subjects_2
    third = second + LMS.number_of_subjects_3
    fourth = third + LMS.number_of_subjects_4
    for row in range(third, fourth):
        total4 = int(ws[f"K{row}"].value)
        total_num_semester_4.append(total4)
    total_sum_4 = sum(total_num_semester_4)
    ws[f"B{last_row + 4}"] = "Semester 4 marks:"
    ws[f"B{last_row + 4}"].font = Font(bold=True, size=11, name="Google Sans")
    ws.cell(row=last_row + 4, column=2).border = LMS.thin_border
    ws[f"C{last_row + 4}"] = total_sum_4
    ws[f"C{last_row + 4}"].font = Font(bold=True, size=11, name="Google Sans")
    ws[f"C{last_row + 4}"].alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=last_row + 4, column=3).border = LMS.thin_border
    # ----- Semester 4 ----- #

    # ----- Semester 5 ----- #
    total_num_semester_5 = []
    first = 5 + LMS.number_of_subjects_1
    second = first + LMS.number_of_subjects_2
    third = second + LMS.number_of_subjects_3
    fourth = third + LMS.number_of_subjects_4
    fifth = fourth + LMS.number_of_subjects_5
    for row in range(fourth, fifth):
        total5 = int(ws[f"K{row}"].value)
        total_num_semester_5.append(total5)
    total_sum_5 = sum(total_num_semester_5)
    ws[f"B{last_row + 5}"] = "Semester 5 marks:"
    ws[f"B{last_row + 5}"].font = Font(bold=True, size=11, name="Google Sans")
    ws.cell(row=last_row + 5, column=2).border = LMS.thin_border
    ws[f"C{last_row + 5}"] = total_sum_5
    ws[f"C{last_row + 5}"].font = Font(bold=True, size=11, name="Google Sans")
    ws[f"C{last_row + 5}"].alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=last_row + 5, column=3).border = LMS.thin_border
    # ----- Semester 5 ----- #

    # ----- Semester 6 ----- #
    total_num_semester_6 = []
    first = 5 + LMS.number_of_subjects_1
    second = first + LMS.number_of_subjects_2
    third = second + LMS.number_of_subjects_3
    fourth = third + LMS.number_of_subjects_4
    fifth = fourth + LMS.number_of_subjects_5
    sixth = fifth + LMS.number_of_subjects_6
    for row in range(fifth, sixth):
        total6 = int(ws[f"K{row}"].value)
        total_num_semester_6.append(total6)
    total_sum_6 = sum(total_num_semester_6)
    ws[f"B{last_row + 6}"] = "Semester 6 marks:"
    ws[f"B{last_row + 6}"].font = Font(bold=True, size=11, name="Google Sans")
    ws.cell(row=last_row + 6, column=2).border = LMS.thin_border
    ws[f"C{last_row + 6}"] = total_sum_6
    ws[f"C{last_row + 6}"].font = Font(bold=True, size=11, name="Google Sans")
    ws[f"C{last_row + 6}"].alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=last_row + 6, column=3).border = LMS.thin_border
    # ----- Semester 6 ---- #

    # ----- Semester 7 ----- #
    total_num_semester_7 = []
    first = 5 + LMS.number_of_subjects_1
    second = first + LMS.number_of_subjects_2
    third = second + LMS.number_of_subjects_3
    fourth = third + LMS.number_of_subjects_4
    fifth = fourth + LMS.number_of_subjects_5
    sixth = fifth + LMS.number_of_subjects_6
    seventh = sixth + LMS.number_of_subjects_7
    for row in range(sixth, seventh):
        total7 = int(ws[f"K{row}"].value)
        total_num_semester_7.append(total7)
    total_sum_7 = sum(total_num_semester_7)
    ws[f"B{last_row + 7}"] = "Semester 7 marks:"
    ws[f"B{last_row + 7}"].font = Font(bold=True, size=11, name="Google Sans")
    ws.cell(row=last_row + 7, column=2).border = LMS.thin_border
    ws[f"C{last_row + 7}"] = total_sum_7
    ws[f"C{last_row + 7}"].font = Font(bold=True, size=11, name="Google Sans")
    ws[f"C{last_row + 7}"].alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=last_row + 7, column=3).border = LMS.thin_border
    # ----- Semester 7 ---- #

    # ----- Semester 8 ----- #
    total_num_semester_8 = []
    first = 5 + LMS.number_of_subjects_1
    second = first + LMS.number_of_subjects_2
    third = second + LMS.number_of_subjects_3
    fourth = third + LMS.number_of_subjects_4
    fifth = fourth + LMS.number_of_subjects_5
    sixth = fifth + LMS.number_of_subjects_6
    seventh = sixth + LMS.number_of_subjects_7
    eight = seventh + LMS.number_of_subjects_8
    for row in range(seventh, eight):
        total8 = int(ws[f"K{row}"].value)
        total_num_semester_8.append(total8)
    total_sum_8 = sum(total_num_semester_8)
    ws[f"B{last_row + 8}"] = "Semester 8 marks:"
    ws[f"B{last_row + 8}"].font = Font(bold=True, size=11, name="Google Sans")
    ws.cell(row=last_row + 8, column=2).border = LMS.thin_border
    ws[f"C{last_row + 8}"] = total_sum_8
    ws[f"C{last_row + 8}"].font = Font(bold=True, size=11, name="Google Sans")
    ws[f"C{last_row + 8}"].alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=last_row + 8, column=3).border = LMS.thin_border
    # ----- Semester 8 ---- #

wb.save(LMS.fname)
path = Path.cwd()
print("File created: {}.xlsx".format(LMS.USERNAME))
print("")
print(f"Locating {LMS.USERNAME}.xlsx")
if os.name == "nt":
    print(f"File located at: {path}/{LMS.USERNAME}.xlsx")
    os.system('explorer.exe /select,"{}.xlsx"'.format(LMS.USERNAME))
elif os.name == "posix":
    print(f"File located at: {path}/{LMS.USERNAME}.xlsx")

print("Killing webdriver process")
print("Done!")

# subprocess.call(r"kill.bat")
go = input("CLose the window to exit")
os._exit(0)






